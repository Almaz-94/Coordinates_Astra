import openpyxl
import xlwt

from openpyxl import load_workbook

garmin = load_workbook('D:\Шамянская площадь\Координаты\Удалени дублей 2 этап\Шамян БЕЗ КОНТРОЛЯ.xlsx')
for sheetsG in garmin:
    maxG = sheetsG.max_row
    for i in range(2,maxG):
        j=i+1
        p=i+2
        if sheetsG.cell(row=i, column=1).value == sheetsG.cell(row=p, column=1).value:

            if abs(float(sheetsG.cell(row=i, column=2).value) - float(sheetsG.cell(row=p, column=2).value)) == 10.0:

                sheetsG.cell(row=i, column=8).value = sheetsG.cell(row=i, column=1).value
                sheetsG.cell(row=i, column=9).value = float(sheetsG.cell(row=i, column=2).value)+5
                sheetsG.cell(row=i, column=10).value = (float(sheetsG.cell(row=i, column=4).value) +
                                                        float(sheetsG.cell(row=p, column=4).value)) / 2
                sheetsG.cell(row=i, column=11).value = (float(sheetsG.cell(row=i, column=5).value) +
                                                        float(sheetsG.cell(row=p, column=5).value)) / 2

            elif abs(float(sheetsG.cell(row=i, column=2).value) - float(sheetsG.cell(row=j, column=2).value)) == 10.0:

                sheetsG.cell(row=i, column=8).value = sheetsG.cell(row=i, column=1).value
                sheetsG.cell(row=i, column=9).value = float(sheetsG.cell(row=i, column=2).value) + 5
                sheetsG.cell(row=i, column=10).value = (float(sheetsG.cell(row=i, column=4).value) +
                                                        float(sheetsG.cell(row=j, column=4).value)) / 2
                sheetsG.cell(row=i, column=11).value = (float(sheetsG.cell(row=i, column=5).value) +
                                                        float(sheetsG.cell(row=j, column=5).value)) / 2
            else:
                sheetsG.cell(row=i, column=8).value='prop'
garmin.save("Шамян С ЦЕНТРОМ ПР ПК БЕЗ КОНТРОЛЯ.xls")

