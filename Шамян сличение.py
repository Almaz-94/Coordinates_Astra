import openpyxl
import xlwt

from openpyxl import load_workbook

baza = load_workbook('D:\Шамянская площадь\Координаты\База Шамян+доп.xlsx')
garmin = load_workbook('D:\Шамянская площадь\Координаты\Гармин Shamyan_Pl_01_80_s_pulkovo.xlsx')
#sh = garmin.get_sheet_by_name

a=baza['gph_200_50_stage']
b=baza['SM_200_50_pirit_WGS84']
c=baza['gph_200_50_stage_3']
for sheetsG in garmin:
        maxG = sheetsG.max_row
        maxB = c.max_row
        for i in range(2, maxG):

            for j in range(2, maxB):

                if abs(float(sheetsG.cell(row=i, column=4).value) - float(c.cell(row=j, column=4).value)) < 8:
                    if abs(float(sheetsG.cell(row=i, column=5).value) - float(c.cell(row=j, column=5).value)) < 8:
                         sheetsG.cell(row=i, column=1).value = c.cell(row=j, column=1).value
                         sheetsG.cell(row=i, column=2).value = c.cell(row=j, column=2).value



garmin.save("Гармин Шамян_3 уч с ПР ПК.xls")