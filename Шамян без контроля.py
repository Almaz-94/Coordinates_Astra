import openpyxl
import xlwt

from openpyxl import load_workbook


garmin = load_workbook('D:\Шамянская площадь\Координаты\Сличение 1 этап\Гармин Шамян сличение.xlsx')

for sheetsG in garmin:
    maxG = sheetsG.max_row

    for i in range(2,maxG-1):
        j=i+1
        if sheetsG.cell(row=i,column=1).value==sheetsG.cell(row=j, column=1).value:

            if sheetsG.cell(row=i,column=2).value==sheetsG.cell(row=j, column=2).value:

                    sheetsG.cell(row=i, column=3).value=99999


garmin.save("Шамян БЕЗ КОНТРОЛЯ.xls")